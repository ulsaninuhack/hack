#!/usr/bin/env python3
"""Aggregate VWorld building-age register records to 156 map geometry zones.

This builder links the normalized VWorld ``AL_D197`` building-age records to
the ``AL_D010`` integrated-building polygons through the official A1 field,
``GIS건물통합식별번호``.  Every integrated polygon is reduced to a Shapely
representative point in EPSG:5186 and spatially assigned to the available
2025-06-30 administrative-dong boundary snapshot.

The output grain is a *2025 geometry zone*, not one of the 162 current
administrative-dong rows.  A normalized building-register record is counted
once only when every GIS identifier listed on that record resolves to the same
single geometry zone.  Missing, outside, overlapping, or cross-zone links are
excluded and fully reconciled in the validation JSON.  Shared GIS identifiers
are expected: when one identifier belongs to multiple normalized register
records, all of those records inherit the identifier's one spatial assignment.

The aggregate is suitable only for local, non-commercial analysis pending a
specific VWorld redistribution/licence review.  It must not be described as a
count of unique buildings, homes, households, or residents.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import json
import math
import shutil
import sys
import tempfile
import time
import warnings
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable


DATA_ROOT = Path(__file__).resolve().parents[1]
LOCAL_DEPS = DATA_ROOT / ".deps"
if LOCAL_DEPS.is_dir():
    sys.path.insert(0, str(LOCAL_DEPS))

try:
    import shapefile  # type: ignore[import-not-found]
    from openpyxl import load_workbook  # type: ignore[import-not-found]
    from pyproj import CRS, Transformer  # type: ignore[import-not-found]
    from shapely import STRtree  # type: ignore[import-not-found]
    from shapely.geometry import (  # type: ignore[import-not-found]
        GeometryCollection,
        MultiPolygon,
        Polygon,
        shape,
    )
    from shapely.ops import transform, unary_union  # type: ignore[import-not-found]
    from shapely.validation import make_valid  # type: ignore[import-not-found]
except ImportError as exc:  # pragma: no cover - dependency failure path
    raise SystemExit(
        "Missing GIS dependencies. Restore work/i5-data/.deps or install "
        f"openpyxl, pyshp, pyproj, and shapely: {exc}"
    ) from exc


DEFAULT_INTEGRATED_ARCHIVE = (
    DATA_ROOT / "raw" / "housing" / "vworld_incheon_building_integrated_20260809.zip"
)
DEFAULT_COLUMN_DICTIONARY = (
    DATA_ROOT / "raw" / "housing" / "vworld_national_key_data_column_dictionary_20260102.xlsx"
)
DEFAULT_AGE_RECORDS = DATA_ROOT / "processed" / "housing_building_age_records.csv.gz"
DEFAULT_AGE_VALIDATION = DATA_ROOT / "processed" / "housing_building_age_validation.json"
DEFAULT_BOUNDARY = DATA_ROOT / "processed" / "boundary_incheon_admin_dong_20250630.geojson"
DEFAULT_BOUNDARY_VALIDATION = DATA_ROOT / "processed" / "boundary_validation.json"
DEFAULT_OUTPUT_DIR = DATA_ROOT / "processed"

SUMMARY_NAME = "housing_admin_geometry_summary_20260805.csv"
VALIDATION_NAME = "housing_admin_geometry_validation_20260805.json"
SCRIPT_VERSION = 1
SOURCE_CRS_EPSG = 5186
BOUNDARY_CRS_EPSG = 4326
EXPECTED_ZONE_COUNT = 156
SOURCE_ENCODING = "cp949"
OUTPUT_ENCODING = "utf-8-sig"


SUMMARY_FIELDS = [
    "geometry_zone_id",
    "geometry_adm_cd_20250630",
    "geometry_district_name_20250630",
    "geometry_dong_name_20250630",
    "geometry_base_date",
    "current_district_names_20260701",
    "current_dong_names_20260701",
    "map_unit_status",
    "integrated_building_geometry_features",
    "integrated_gis_ids_with_feature_in_zone",
    "linked_gis_ids_used_by_assigned_age_records",
    "age_building_register_records_assigned",
    "age_valid_records",
    "age_missing_or_invalid_records",
    "age_coverage_pct",
    "age_30_plus_records",
    "age_30_plus_share_valid_pct",
    "age_40_plus_records",
    "age_40_plus_share_valid_pct",
    "residential_records",
    "residential_age_valid_records",
    "residential_age_missing_or_invalid_records",
    "residential_age_coverage_pct",
    "residential_age_30_plus_records",
    "residential_age_30_plus_share_valid_pct",
    "residential_age_40_plus_records",
    "residential_age_40_plus_share_valid_pct",
    "metric_grain",
    "spatial_assignment_method",
    "age_data_as_of_date",
]


@dataclass(frozen=True)
class AgeRecord:
    """Only the normalized fields required for a non-expanding spatial join."""

    record_key: str
    gis_ids: tuple[str, ...]
    age: int | None
    age_valid: bool
    residential: bool
    source_row_count: int


@dataclass
class ReferencedGeometryState:
    """Spatial evidence observed for one age-referenced GIS identifier."""

    feature_count: int = 0
    zones: set[str] | None = None
    outside_feature_count: int = 0
    overlapping_feature_count: int = 0
    invalid_or_empty_feature_count: int = 0

    def __post_init__(self) -> None:
        if self.zones is None:
            self.zones = set()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--integrated-archive", type=Path, default=DEFAULT_INTEGRATED_ARCHIVE)
    parser.add_argument("--column-dictionary", type=Path, default=DEFAULT_COLUMN_DICTIONARY)
    parser.add_argument("--age-records", type=Path, default=DEFAULT_AGE_RECORDS)
    parser.add_argument("--age-validation", type=Path, default=DEFAULT_AGE_VALIDATION)
    parser.add_argument("--boundary", type=Path, default=DEFAULT_BOUNDARY)
    parser.add_argument("--boundary-validation", type=Path, default=DEFAULT_BOUNDARY_VALIDATION)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def atomic_write_csv(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp")
    with temporary.open("w", encoding=OUTPUT_ENCODING, newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SUMMARY_FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    temporary.replace(path)


def atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp")
    with temporary.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2, sort_keys=True)
        handle.write("\n")
    temporary.replace(path)


def percentage(numerator: int, denominator: int) -> str:
    if denominator == 0:
        return ""
    return f"{numerator / denominator * 100:.4f}"


def parse_integer(value: str) -> int | None:
    text = (value or "").strip()
    if not text:
        return None
    try:
        return int(text)
    except ValueError as exc:
        raise ValueError(f"Expected integer, got {value!r}") from exc


def parse_bool(value: str, field: str) -> bool:
    normalized = (value or "").strip().lower()
    if normalized == "true":
        return True
    if normalized == "false":
        return False
    raise ValueError(f"Expected true/false in {field}, got {value!r}")


def polygonal_only(geometry: Any) -> Any:
    if isinstance(geometry, (Polygon, MultiPolygon)):
        return geometry
    if isinstance(geometry, GeometryCollection):
        polygonal_parts: list[Any] = []
        for part in geometry.geoms:
            if isinstance(part, Polygon):
                polygonal_parts.append(part)
            elif isinstance(part, MultiPolygon):
                polygonal_parts.extend(part.geoms)
        if polygonal_parts:
            return unary_union(polygonal_parts)
    raise ValueError(f"Geometry has no polygonal component: {geometry.geom_type}")


def verify_field_dictionary(path: Path) -> dict[str, Any]:
    """Verify the official AL_D010 A1 meaning rather than assuming it."""

    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["테이블정의서(전체)"]
    fields: dict[str, dict[str, Any]] = {}
    for excel_row, values in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = tuple(values)
        if len(values) < 10 or str(values[3] or "").strip() != "AL_D010":
            continue
        field_name = str(values[6] or "").strip()
        fields[field_name] = {
            "excel_row": excel_row,
            "korean_name": str(values[7] or "").strip(),
            "required": str(values[9] or "").strip(),
        }
    workbook.close()
    if fields.get("A1", {}).get("korean_name") != "GIS건물통합식별번호":
        raise ValueError(f"Official dictionary does not map AL_D010 A1 as expected: {fields.get('A1')}")
    if fields.get("A22", {}).get("korean_name") != "데이터기준일자":
        raise ValueError(f"Official dictionary does not map AL_D010 A22 as expected: {fields.get('A22')}")
    return {
        "workbook": path.name,
        "workbook_sha256": sha256_file(path),
        "sheet": sheet.title,
        "table": "GIS건물통합정보_AL_D010",
        "verified_fields": {key: fields[key] for key in ("A1", "A22")},
    }


def load_age_records(path: Path) -> tuple[list[AgeRecord], Counter[str], dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(path)
    required = {
        "building_key",
        "gis_building_ids",
        "building_age_official",
        "building_age_valid",
        "is_residential",
        "source_row_count",
    }
    records: list[AgeRecord] = []
    gis_id_reference_counts: Counter[str] = Counter()
    record_keys: set[str] = set()
    records_with_multiple_ids = 0
    source_rows_represented = 0
    data_dates: Counter[str] = Counter()
    with gzip.open(path, "rt", encoding=OUTPUT_ENCODING, newline="") as handle:
        reader = csv.DictReader(handle)
        missing = required.difference(reader.fieldnames or [])
        if missing:
            raise ValueError(f"Age records missing required fields: {sorted(missing)}")
        for row in reader:
            key = row["building_key"].strip()
            if not key or key in record_keys:
                raise ValueError(f"Blank or duplicate normalized building_key: {key!r}")
            record_keys.add(key)
            gis_ids = tuple(sorted({item.strip() for item in row["gis_building_ids"].split("|") if item.strip()}))
            if not gis_ids:
                raise ValueError(f"Normalized record lacks GIS identifier: {key}")
            if len(gis_ids) > 1:
                records_with_multiple_ids += 1
            gis_id_reference_counts.update(gis_ids)
            age = parse_integer(row["building_age_official"])
            age_valid = parse_bool(row["building_age_valid"], "building_age_valid")
            if age_valid and (age is None or not 0 <= age <= 200):
                raise ValueError(f"Valid age flag contradicts age for {key}: {age!r}")
            residential = parse_bool(row["is_residential"], "is_residential")
            source_row_count = parse_integer(row["source_row_count"])
            if source_row_count is None or source_row_count < 1:
                raise ValueError(f"Invalid source_row_count for {key}")
            source_rows_represented += source_row_count
            data_date = row.get("data_as_of_date", "").strip()
            if data_date:
                data_dates[data_date] += 1
            records.append(
                AgeRecord(
                    record_key=key,
                    gis_ids=gis_ids,
                    age=age,
                    age_valid=age_valid,
                    residential=residential,
                    source_row_count=source_row_count,
                )
            )
    records.sort(key=lambda record: record.record_key)
    diagnostics = {
        "normalized_record_count": len(records),
        "normalized_record_keys_unique": len(record_keys) == len(records),
        "source_rows_represented": source_rows_represented,
        "records_with_multiple_gis_ids": records_with_multiple_ids,
        "gis_id_references": sum(gis_id_reference_counts.values()),
        "unique_gis_ids": len(gis_id_reference_counts),
        "gis_ids_shared_by_multiple_records": sum(
            count > 1 for count in gis_id_reference_counts.values()
        ),
        "maximum_records_per_gis_id": max(gis_id_reference_counts.values(), default=0),
        "data_as_of_dates": dict(sorted(data_dates.items())),
    }
    return records, gis_id_reference_counts, diagnostics


def load_boundaries(path: Path) -> tuple[list[dict[str, Any]], list[Any], STRtree, dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(path)
    payload = json.loads(path.read_text(encoding="utf-8"))
    transformer = Transformer.from_crs(
        f"EPSG:{BOUNDARY_CRS_EPSG}", f"EPSG:{SOURCE_CRS_EPSG}", always_xy=True
    )
    zones: list[dict[str, Any]] = []
    polygons: list[Any] = []
    repaired = 0
    for feature in payload.get("features", []):
        properties = feature.get("properties", {})
        zone_id = str(properties.get("geometry_zone_id", "")).strip()
        if not zone_id:
            raise ValueError("Boundary feature lacks geometry_zone_id")
        geometry = shape(feature["geometry"])
        if geometry.is_empty:
            raise ValueError(f"Empty boundary geometry: {zone_id}")
        geometry = transform(transformer.transform, geometry)
        if not geometry.is_valid:
            geometry = polygonal_only(make_valid(geometry))
            repaired += 1
        zones.append(properties)
        polygons.append(geometry)
    zone_ids = [str(zone["geometry_zone_id"]) for zone in zones]
    if len(zones) != EXPECTED_ZONE_COUNT or len(set(zone_ids)) != len(zone_ids):
        raise ValueError(
            f"Expected {EXPECTED_ZONE_COUNT} unique boundary zones, got {len(zones)} / {len(set(zone_ids))}"
        )
    return zones, polygons, STRtree(polygons), {
        "feature_count": len(zones),
        "unique_geometry_zone_ids": len(set(zone_ids)),
        "boundary_repairs_after_projection": repaired,
        "source_crs": f"EPSG:{BOUNDARY_CRS_EPSG}",
        "analysis_crs": f"EPSG:{SOURCE_CRS_EPSG}",
        "is_simplified_2025_snapshot": True,
    }


def safe_extract_shapefile(archive_path: Path, temporary_dir: Path) -> dict[str, Path]:
    """Extract only the four required, flat shapefile members."""

    required_suffixes = {".shp", ".shx", ".dbf", ".prj"}
    extracted: dict[str, Path] = {}
    with zipfile.ZipFile(archive_path) as archive:
        for member in archive.infolist():
            suffix = Path(member.filename).suffix.lower()
            if suffix not in required_suffixes:
                continue
            if Path(member.filename).name != member.filename or member.is_dir():
                raise ValueError(f"Unsafe or nested shapefile member: {member.filename!r}")
            if suffix in extracted:
                raise ValueError(f"Duplicate {suffix} member in {archive_path.name}")
            target = temporary_dir / member.filename
            with archive.open(member) as source, target.open("wb") as destination:
                shutil.copyfileobj(source, destination, length=1024 * 1024)
            extracted[suffix] = target
    missing = required_suffixes.difference(extracted)
    if missing:
        raise ValueError(f"Integrated archive missing members: {sorted(missing)}")
    return extracted


def verify_integrated_crs(prj_path: Path) -> tuple[CRS, str]:
    wkt = prj_path.read_text(encoding="ascii", errors="strict").strip()
    crs = CRS.from_wkt(wkt)
    # This ESRI WKT names the datum differently from the current EPSG record,
    # so PROJ gives the otherwise exact match less than 50% confidence.  The
    # source itself declares EPSG:5186; also verify every material projection
    # parameter before accepting that declaration.
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=r"You will likely lose important projection information.*",
        )
        parameters = crs.to_dict()
    declared = 'AUTHORITY["EPSG","5186"]' in wkt.replace(" ", "")
    expected_parameters = {
        "proj": "tmerc",
        "lat_0": 38,
        "lon_0": 127,
        "k": 1,
        "x_0": 200000,
        "y_0": 600000,
        "ellps": "GRS80",
        "units": "m",
    }
    parameters_match = all(parameters.get(key) == value for key, value in expected_parameters.items())
    if not declared or not parameters_match:
        raise ValueError(
            f"Expected EPSG:{SOURCE_CRS_EPSG}; declared={declared}, parameters={parameters}"
        )
    return crs, wkt


def assign_representative_point(point: Any, tree: STRtree, polygons: list[Any]) -> list[int]:
    candidates = tree.query(point)
    return [int(index) for index in candidates if polygons[int(index)].covers(point)]


def spatialize_integrated_buildings(
    archive_path: Path,
    referenced_age_ids: set[str],
    zones: list[dict[str, Any]],
    polygons: list[Any],
    tree: STRtree,
) -> tuple[
    dict[str, ReferencedGeometryState],
    Counter[str],
    dict[str, set[str]],
    Counter[str],
    dict[str, Any],
]:
    if not archive_path.exists():
        raise FileNotFoundError(archive_path)
    referenced_states: dict[str, ReferencedGeometryState] = {}
    feature_counts_by_zone: Counter[str] = Counter()
    gis_ids_with_feature_by_zone: dict[str, set[str]] = defaultdict(set)
    all_gis_id_counts: Counter[str] = Counter()
    diagnostics: Counter[str] = Counter()
    data_dates: Counter[str] = Counter()
    shape_types: Counter[str] = Counter()
    repaired_features = 0
    shapefile_fields: list[str] = []
    projection_wkt = ""

    with tempfile.TemporaryDirectory(prefix="i5_vworld_integrated_") as temporary_name:
        extracted = safe_extract_shapefile(archive_path, Path(temporary_name))
        _, projection_wkt = verify_integrated_crs(extracted[".prj"])
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message=r"Specified encoding: cp949 different.*")
            reader = shapefile.Reader(str(extracted[".shp"]), encoding=SOURCE_ENCODING)
            shapefile_fields = [field.name for field in reader.fields[1:]]
            if "A1" not in shapefile_fields or "A22" not in shapefile_fields:
                raise ValueError(f"Integrated DBF lacks A1/A22: {shapefile_fields}")
            if reader.numRecords != reader.numShapes:
                raise ValueError("Integrated SHP/DBF record-count mismatch")
            for shape_record in reader.iterShapeRecords(fields=["A1", "A22"]):
                diagnostics["shape_attribute_records"] += 1
                gis_id = str(shape_record.record[0] or "").strip()
                data_date = str(shape_record.record[1] or "").strip()
                if data_date:
                    data_dates[data_date] += 1
                if not gis_id:
                    diagnostics["blank_gis_id_features"] += 1
                else:
                    all_gis_id_counts[gis_id] += 1
                state = None
                if gis_id in referenced_age_ids:
                    state = referenced_states.setdefault(gis_id, ReferencedGeometryState())
                    state.feature_count += 1

                source_shape = shape_record.shape
                shape_types[str(source_shape.shapeType)] += 1
                if source_shape.shapeType not in {5, 15, 25, 31}:
                    diagnostics["unsupported_geometry_features"] += 1
                    if state is not None:
                        state.invalid_or_empty_feature_count += 1
                    continue
                try:
                    geometry = shape(source_shape.__geo_interface__)
                    if geometry.is_empty:
                        raise ValueError("empty geometry")
                    if not geometry.is_valid:
                        geometry = polygonal_only(make_valid(geometry))
                        repaired_features += 1
                    representative_point = geometry.representative_point()
                    matches = assign_representative_point(representative_point, tree, polygons)
                except Exception:
                    diagnostics["invalid_or_empty_geometry_features"] += 1
                    if state is not None:
                        state.invalid_or_empty_feature_count += 1
                    continue

                if not matches:
                    diagnostics["representative_points_outside_boundaries"] += 1
                    if state is not None:
                        state.outside_feature_count += 1
                    continue
                if len(matches) > 1:
                    diagnostics["representative_points_in_overlapping_boundaries"] += 1
                    if state is not None:
                        state.overlapping_feature_count += 1
                    continue

                zone_id = str(zones[matches[0]]["geometry_zone_id"])
                feature_counts_by_zone[zone_id] += 1
                if gis_id:
                    gis_ids_with_feature_by_zone[zone_id].add(gis_id)
                if state is not None:
                    assert state.zones is not None
                    state.zones.add(zone_id)

    diagnostics_payload = {
        "archive": archive_path.name,
        "archive_sha256": sha256_file(archive_path),
        "archive_bytes": archive_path.stat().st_size,
        "shapefile_fields": shapefile_fields,
        "projection_wkt": projection_wkt,
        "source_crs": f"EPSG:{SOURCE_CRS_EPSG}",
        "shape_type_counts": dict(sorted(shape_types.items())),
        "shape_attribute_records": diagnostics["shape_attribute_records"],
        "blank_gis_id_features": diagnostics["blank_gis_id_features"],
        "unique_nonblank_gis_ids": len(all_gis_id_counts),
        "gis_ids_with_multiple_geometry_features": sum(
            count > 1 for count in all_gis_id_counts.values()
        ),
        "maximum_geometry_features_per_gis_id": max(all_gis_id_counts.values(), default=0),
        "geometry_features_repaired": repaired_features,
        "invalid_or_empty_geometry_features": diagnostics["invalid_or_empty_geometry_features"],
        "unsupported_geometry_features": diagnostics["unsupported_geometry_features"],
        "representative_points_assigned_once": sum(feature_counts_by_zone.values()),
        "representative_points_outside_boundaries": diagnostics[
            "representative_points_outside_boundaries"
        ],
        "representative_points_in_overlapping_boundaries": diagnostics[
            "representative_points_in_overlapping_boundaries"
        ],
        "data_as_of_dates": dict(sorted(data_dates.items())),
    }
    return (
        referenced_states,
        feature_counts_by_zone,
        gis_ids_with_feature_by_zone,
        all_gis_id_counts,
        diagnostics_payload,
    )


def classify_referenced_gis_ids(
    referenced_ids: set[str], states: dict[str, ReferencedGeometryState]
) -> tuple[dict[str, str], dict[str, Any]]:
    resolved: dict[str, str] = {}
    reasons: Counter[str] = Counter()
    ids_with_multiple_features = 0
    for gis_id in sorted(referenced_ids):
        state = states.get(gis_id)
        if state is None:
            reasons["no_integrated_geometry"] += 1
            continue
        if state.feature_count > 1:
            ids_with_multiple_features += 1
        if state.invalid_or_empty_feature_count:
            reasons["invalid_or_empty_geometry_feature"] += 1
            continue
        if state.outside_feature_count:
            reasons["feature_outside_156_zone_boundary"] += 1
            continue
        if state.overlapping_feature_count:
            reasons["feature_in_overlapping_boundaries"] += 1
            continue
        zones = state.zones or set()
        if len(zones) == 0:
            reasons["no_resolved_zone"] += 1
            continue
        if len(zones) > 1:
            reasons["geometry_features_span_multiple_zones"] += 1
            continue
        resolved[gis_id] = next(iter(zones))
    return resolved, {
        "referenced_gis_ids": len(referenced_ids),
        "resolved_to_exactly_one_zone": len(resolved),
        "excluded_by_reason": dict(sorted(reasons.items())),
        "referenced_ids_with_multiple_geometry_features": ids_with_multiple_features,
    }


def aggregate_age_records(
    records: list[AgeRecord],
    resolved_gis_zones: dict[str, str],
    gis_id_reference_counts: Counter[str],
) -> tuple[dict[str, Counter[str]], dict[str, set[str]], dict[str, Any]]:
    stats_by_zone: dict[str, Counter[str]] = defaultdict(Counter)
    linked_ids_by_zone: dict[str, set[str]] = defaultdict(set)
    exclusion_reasons: Counter[str] = Counter()
    assigned_keys: set[str] = set()
    assigned_records_using_shared_id = 0
    assigned_records_with_multiple_ids = 0
    assigned_source_rows_represented = 0

    for record in records:
        missing_ids = [gis_id for gis_id in record.gis_ids if gis_id not in resolved_gis_zones]
        if missing_ids:
            exclusion_reasons["one_or_more_gis_ids_not_single_zone"] += 1
            continue
        zones = {resolved_gis_zones[gis_id] for gis_id in record.gis_ids}
        if len(zones) != 1:
            exclusion_reasons["multiple_gis_ids_disagree_on_zone"] += 1
            continue
        zone_id = next(iter(zones))
        if record.record_key in assigned_keys:
            raise AssertionError(f"Normalized record double-counted: {record.record_key}")
        assigned_keys.add(record.record_key)
        linked_ids_by_zone[zone_id].update(record.gis_ids)
        assigned_source_rows_represented += record.source_row_count
        if len(record.gis_ids) > 1:
            assigned_records_with_multiple_ids += 1
        if any(gis_id_reference_counts[gis_id] > 1 for gis_id in record.gis_ids):
            assigned_records_using_shared_id += 1

        stats = stats_by_zone[zone_id]
        stats["age_building_register_records_assigned"] += 1
        if record.age_valid:
            stats["age_valid_records"] += 1
            assert record.age is not None
            if record.age >= 30:
                stats["age_30_plus_records"] += 1
            if record.age >= 40:
                stats["age_40_plus_records"] += 1
        else:
            stats["age_missing_or_invalid_records"] += 1

        if record.residential:
            stats["residential_records"] += 1
            if record.age_valid:
                stats["residential_age_valid_records"] += 1
                assert record.age is not None
                if record.age >= 30:
                    stats["residential_age_30_plus_records"] += 1
                if record.age >= 40:
                    stats["residential_age_40_plus_records"] += 1
            else:
                stats["residential_age_missing_or_invalid_records"] += 1

    return stats_by_zone, linked_ids_by_zone, {
        "source_normalized_records": len(records),
        "assigned_normalized_records": len(assigned_keys),
        "excluded_normalized_records": len(records) - len(assigned_keys),
        "excluded_record_reasons": dict(sorted(exclusion_reasons.items())),
        "assigned_record_keys_unique": len(assigned_keys),
        "assigned_records_with_multiple_gis_ids": assigned_records_with_multiple_ids,
        "assigned_records_using_gis_id_shared_by_multiple_records": assigned_records_using_shared_id,
        "assigned_source_rows_represented": assigned_source_rows_represented,
    }


def build_summary_rows(
    zones: list[dict[str, Any]],
    feature_counts_by_zone: Counter[str],
    gis_ids_with_feature_by_zone: dict[str, set[str]],
    linked_ids_by_zone: dict[str, set[str]],
    stats_by_zone: dict[str, Counter[str]],
    age_data_dates: list[str],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for zone in sorted(zones, key=lambda item: str(item["geometry_adm_cd_20250630"])):
        zone_id = str(zone["geometry_zone_id"])
        stats = stats_by_zone[zone_id]
        records = stats["age_building_register_records_assigned"]
        valid = stats["age_valid_records"]
        residential = stats["residential_records"]
        residential_valid = stats["residential_age_valid_records"]
        rows.append(
            {
                "geometry_zone_id": zone_id,
                "geometry_adm_cd_20250630": str(zone["geometry_adm_cd_20250630"]),
                "geometry_district_name_20250630": str(zone["geometry_district_name_20250630"]),
                "geometry_dong_name_20250630": str(zone["geometry_dong_name_20250630"]),
                "geometry_base_date": str(zone["geometry_base_date"]),
                "current_district_names_20260701": "|".join(zone["current_district_names_20260701"]),
                "current_dong_names_20260701": "|".join(zone["current_dong_names_20260701"]),
                "map_unit_status": str(zone["map_unit_status"]),
                "integrated_building_geometry_features": str(feature_counts_by_zone[zone_id]),
                "integrated_gis_ids_with_feature_in_zone": str(
                    len(gis_ids_with_feature_by_zone.get(zone_id, set()))
                ),
                "linked_gis_ids_used_by_assigned_age_records": str(
                    len(linked_ids_by_zone.get(zone_id, set()))
                ),
                "age_building_register_records_assigned": str(records),
                "age_valid_records": str(valid),
                "age_missing_or_invalid_records": str(stats["age_missing_or_invalid_records"]),
                "age_coverage_pct": percentage(valid, records),
                "age_30_plus_records": str(stats["age_30_plus_records"]),
                "age_30_plus_share_valid_pct": percentage(stats["age_30_plus_records"], valid),
                "age_40_plus_records": str(stats["age_40_plus_records"]),
                "age_40_plus_share_valid_pct": percentage(stats["age_40_plus_records"], valid),
                "residential_records": str(residential),
                "residential_age_valid_records": str(residential_valid),
                "residential_age_missing_or_invalid_records": str(
                    stats["residential_age_missing_or_invalid_records"]
                ),
                "residential_age_coverage_pct": percentage(residential_valid, residential),
                "residential_age_30_plus_records": str(
                    stats["residential_age_30_plus_records"]
                ),
                "residential_age_30_plus_share_valid_pct": percentage(
                    stats["residential_age_30_plus_records"], residential_valid
                ),
                "residential_age_40_plus_records": str(
                    stats["residential_age_40_plus_records"]
                ),
                "residential_age_40_plus_share_valid_pct": percentage(
                    stats["residential_age_40_plus_records"], residential_valid
                ),
                "metric_grain": "normalized_building_register_record",
                "spatial_assignment_method": "AL_D010_A1_representative_point_within_20250630_zone",
                "age_data_as_of_date": "|".join(age_data_dates),
            }
        )
    return rows


def sum_summary(rows: list[dict[str, str]], field: str) -> int:
    return sum(int(row[field] or 0) for row in rows)


def load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(path)
    return json.loads(path.read_text(encoding="utf-8"))


def main() -> None:
    started = time.perf_counter()
    args = parse_args()
    output_dir = args.output_dir.resolve()
    summary_path = output_dir / SUMMARY_NAME
    validation_path = output_dir / VALIDATION_NAME

    dictionary_validation = verify_field_dictionary(args.column_dictionary.resolve())
    age_records, gis_id_reference_counts, age_diagnostics = load_age_records(
        args.age_records.resolve()
    )
    age_validation = load_json(args.age_validation.resolve())
    boundary_validation = load_json(args.boundary_validation.resolve())
    zones, polygons, tree, boundary_diagnostics = load_boundaries(args.boundary.resolve())
    (
        referenced_states,
        feature_counts_by_zone,
        gis_ids_with_feature_by_zone,
        all_gis_id_counts,
        geometry_diagnostics,
    ) = spatialize_integrated_buildings(
        args.integrated_archive.resolve(),
        set(gis_id_reference_counts),
        zones,
        polygons,
        tree,
    )
    resolved_gis_zones, gis_linkage_diagnostics = classify_referenced_gis_ids(
        set(gis_id_reference_counts), referenced_states
    )
    stats_by_zone, linked_ids_by_zone, record_linkage_diagnostics = aggregate_age_records(
        age_records, resolved_gis_zones, gis_id_reference_counts
    )
    age_data_dates = sorted(age_diagnostics["data_as_of_dates"])
    summary_rows = build_summary_rows(
        zones,
        feature_counts_by_zone,
        gis_ids_with_feature_by_zone,
        linked_ids_by_zone,
        stats_by_zone,
        age_data_dates,
    )
    atomic_write_csv(summary_path, summary_rows)

    assigned = sum_summary(summary_rows, "age_building_register_records_assigned")
    valid = sum_summary(summary_rows, "age_valid_records")
    old_30 = sum_summary(summary_rows, "age_30_plus_records")
    old_40 = sum_summary(summary_rows, "age_40_plus_records")
    residential = sum_summary(summary_rows, "residential_records")
    residential_valid = sum_summary(summary_rows, "residential_age_valid_records")
    residential_old_30 = sum_summary(summary_rows, "residential_age_30_plus_records")
    residential_old_40 = sum_summary(summary_rows, "residential_age_40_plus_records")

    upstream_totals = age_validation["outputs"]["totals"]
    checks = {
        "official_dictionary_maps_AL_D010_A1_to_gis_building_id": True,
        "official_dictionary_maps_AL_D010_A22_to_data_date": True,
        "boundary_zone_count_is_156": len(summary_rows) == EXPECTED_ZONE_COUNT,
        "boundary_validation_status_accepted": boundary_validation.get("status") == "pass",
        "normalized_age_record_count_matches_upstream": len(age_records)
        == int(upstream_totals["building_records"]),
        "normalized_source_row_count_matches_upstream": age_diagnostics["source_rows_represented"]
        == int(upstream_totals["raw_source_rows"]),
        "integrated_shape_attribute_count_matches_upstream": geometry_diagnostics[
            "shape_attribute_records"
        ]
        == int(age_validation["integrated_building_shapefile"]["dbf_record_count"]),
        "every_integrated_feature_reconciled": geometry_diagnostics[
            "shape_attribute_records"
        ]
        == geometry_diagnostics["representative_points_assigned_once"]
        + geometry_diagnostics["representative_points_outside_boundaries"]
        + geometry_diagnostics["representative_points_in_overlapping_boundaries"]
        + geometry_diagnostics["invalid_or_empty_geometry_features"]
        + geometry_diagnostics["unsupported_geometry_features"],
        "assigned_plus_excluded_equals_normalized_records": assigned
        + record_linkage_diagnostics["excluded_normalized_records"]
        == len(age_records),
        "summary_total_matches_assignment_total": assigned
        == record_linkage_diagnostics["assigned_normalized_records"],
        "assigned_records_not_double_counted": assigned
        == record_linkage_diagnostics["assigned_record_keys_unique"],
        "age_valid_plus_invalid_equals_assigned": valid
        + sum_summary(summary_rows, "age_missing_or_invalid_records")
        == assigned,
        "residential_valid_plus_invalid_equals_residential": residential_valid
        + sum_summary(summary_rows, "residential_age_missing_or_invalid_records")
        == residential,
        "all_30_plus_counts_within_valid_counts": old_30 <= valid
        and residential_old_30 <= residential_valid,
        "all_40_plus_counts_within_30_plus_counts": old_40 <= old_30
        and residential_old_40 <= residential_old_30,
        "all_residential_counts_within_all_records": residential <= assigned,
        "summary_has_no_nonfinite_shares": all(
            not value or math.isfinite(float(value))
            for row in summary_rows
            for key, value in row.items()
            if key.endswith("_pct")
        ),
        # The minimum is a pipeline health guard, not a claim that excluded
        # records are ignorable. Exact exclusions and the observed rate remain
        # first-class output fields below.
        "strict_linkage_coverage_at_least_95_percent": assigned / len(age_records) >= 0.95,
    }
    failed_checks = sorted(key for key, passed in checks.items() if not passed)
    if failed_checks:
        raise AssertionError(f"Housing spatial validation failed: {failed_checks}")

    validation = {
        "status": "pass_with_documented_exclusions",
        "script_version": SCRIPT_VERSION,
        "inputs": {
            "integrated_building_archive": str(args.integrated_archive.resolve()),
            "integrated_building_archive_sha256": sha256_file(args.integrated_archive.resolve()),
            "age_records": str(args.age_records.resolve()),
            "age_records_sha256": sha256_file(args.age_records.resolve()),
            "age_validation": str(args.age_validation.resolve()),
            "boundary_geojson": str(args.boundary.resolve()),
            "boundary_geojson_sha256": sha256_file(args.boundary.resolve()),
            "boundary_validation": str(args.boundary_validation.resolve()),
            "column_dictionary": str(args.column_dictionary.resolve()),
        },
        "official_field_dictionary": dictionary_validation,
        "boundary": boundary_diagnostics,
        "normalized_age_records": age_diagnostics,
        "integrated_building_geometry": geometry_diagnostics,
        "gis_identifier_linkage": {
            **gis_linkage_diagnostics,
            "gis_ids_shared_by_multiple_normalized_records": sum(
                count > 1 for count in gis_id_reference_counts.values()
            ),
            "gis_ids_with_multiple_integrated_geometry_features_all": sum(
                count > 1 for count in all_gis_id_counts.values()
            ),
            "rule": (
                "A GIS ID is usable only when every associated integrated-building feature "
                "has a representative point in the same one boundary zone and none is "
                "missing, outside, overlapping, invalid, or empty."
            ),
        },
        "record_assignment": {
            **record_linkage_diagnostics,
            "rule": (
                "A normalized register record is counted once only when all of its listed "
                "GIS IDs resolve and agree on one zone. Multiple records sharing one GIS ID "
                "each inherit that same zone; they are not collapsed because the metric "
                "grain is the normalized building-register record."
            ),
            "strict_assignment_coverage_pct": round(assigned / len(age_records) * 100, 6),
        },
        "outputs": {
            "summary": str(summary_path),
            "summary_sha256": sha256_file(summary_path),
            "summary_rows": len(summary_rows),
            "city_totals": {
                "integrated_building_geometry_features_assigned_to_zones": sum_summary(
                    summary_rows, "integrated_building_geometry_features"
                ),
                "age_building_register_records_assigned": assigned,
                "age_valid_records": valid,
                "age_missing_or_invalid_records": sum_summary(
                    summary_rows, "age_missing_or_invalid_records"
                ),
                "age_30_plus_records": old_30,
                "age_40_plus_records": old_40,
                "residential_records": residential,
                "residential_age_valid_records": residential_valid,
                "residential_age_missing_or_invalid_records": sum_summary(
                    summary_rows, "residential_age_missing_or_invalid_records"
                ),
                "residential_age_30_plus_records": residential_old_30,
                "residential_age_40_plus_records": residential_old_40,
            },
        },
        "checks": checks,
        "failed_checks": failed_checks,
        "metric_semantics": {
            "grain": "normalized VWorld building-register record",
            "not_equivalent_to": [
                "unique physical building",
                "dwelling unit",
                "household",
                "resident",
                "person needing or missing care",
            ],
            "residential_definition_inherited_from_upstream": (
                "is_residential=true only when build_housing.py classified the source main-use "
                "code as 01000 (detached house) or 02000 (multi-unit housing)"
            ),
            "old_housing_thresholds_years": [30, 40],
            "old_share_denominator": "assigned register records with a valid official building age",
            "spatial_unit": (
                "156 VWorld/SGIS 2025-06-30 geometry zones; no 2026 polygon split is fabricated"
            ),
        },
        "limitations": [
            "The boundary GeoJSON is a 10-metre-simplified 2025-06-30 snapshot, while building geometry and age data are from August 2026.",
            "Strictly unresolved GIS IDs and records are excluded rather than nearest-neighbour filled or assigned by legal-dong name.",
            "One GIS ID may refer to multiple normalized register records and one GIS ID may have multiple integrated polygon features; the validation reports both collision types.",
            "The residential subset inherits the narrow upstream main-use-code definition and is not a count of homes or housing units.",
            "No per-building coordinates or geometries are emitted; only 156-zone aggregates are written.",
        ],
        "licence_and_redistribution": {
            "review_required_before_publication": True,
            "intended_use": "local non-commercial hackathon analysis",
            "reason": (
                "VWorld material in this data pack carries a CC BY-NC-ND caution. This output "
                "is an aggregate derived by filtering, joining, and spatially transforming "
                "VWorld sources, so public redistribution must be reviewed before use."
            ),
            "raw_sources_should_not_be_republished": True,
            "this_script_is_not_legal_advice": True,
        },
    }
    atomic_write_json(validation_path, validation)
    elapsed = time.perf_counter() - started
    print(
        json.dumps(
            {
                "summary": str(summary_path),
                "validation": str(validation_path),
                "zones": len(summary_rows),
                "assigned_register_records": assigned,
                "excluded_register_records": record_linkage_diagnostics[
                    "excluded_normalized_records"
                ],
                "coverage_pct": round(assigned / len(age_records) * 100, 6),
                "runtime_seconds": round(elapsed, 3),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
