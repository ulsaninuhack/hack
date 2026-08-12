#!/usr/bin/env python3
"""Build the authoritative 2026-07-01 Incheon administrative-code tables.

This deliberately does not fabricate geometry.  It extracts current MOIS codes and
builds a deterministic name-based bridge from the 2026-06 population snapshot to
the 2026-07 administrative system.
"""

from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "raw"
OUT = ROOT / "processed"


def clean_name(value: str) -> str:
    value = (value or "").strip()
    return (
        value.replace("·", ".")
        .replace(",", ".")
        .replace(" ", "")
        .replace("3.5동", "3.5동")
    )


def load_current_codes() -> list[dict[str, str]]:
    path = RAW / "boundaries" / "mois_admin_dong_codes_20260701.xlsx"
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for code, sido, district, dong, created, deleted in ws.iter_rows(min_row=2, values_only=True):
        if sido != "인천광역시" or not dong or deleted:
            continue
        rows.append(
            {
                "admin_dong_code_20260701": str(code),
                "sido_name_20260701": sido,
                "district_name_20260701": district,
                "dong_name_20260701": dong,
                "created_date": str(created),
            }
        )
    return rows


def load_pre_reform_keys() -> list[dict[str, str]]:
    path = RAW / "demographics" / "korea_population_by_sex_age_admin_dong_20260630.csv"
    with path.open(encoding="cp949", newline="") as handle:
        return [row for row in csv.DictReader(handle) if row["시도명"] == "인천광역시"]


def main() -> None:
    current = load_current_codes()
    previous = load_pre_reform_keys()

    assert len(current) == 162, f"expected 162 current Incheon dongs, got {len(current)}"
    assert len(previous) == 162, f"expected 162 June Incheon dongs, got {len(previous)}"

    by_name: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in current:
        by_name[clean_name(row["dong_name_20260701"])].append(row)

    bridge = []
    unmatched = []
    for old in previous:
        key = clean_name(old["읍면동명"])
        candidates = by_name.get(key, [])
        if len(candidates) != 1:
            unmatched.append({"old": old, "candidate_count": len(candidates)})
            continue
        new = candidates[0]
        bridge.append(
            {
                "source_date": old["기준연월"],
                "old_admin_dong_code": old["행정기관코드"],
                "old_district_name": old["시군구명"],
                "old_dong_name": old["읍면동명"],
                "admin_dong_code_20260701": new["admin_dong_code_20260701"],
                "district_name_20260701": new["district_name_20260701"],
                "dong_name_20260701": new["dong_name_20260701"],
                "mapping_method": "unique_normalized_dong_name",
            }
        )

    if unmatched:
        raise RuntimeError(f"unmatched or ambiguous June dongs: {unmatched}")
    if len({r["admin_dong_code_20260701"] for r in bridge}) != 162:
        raise RuntimeError("bridge is not one-to-one")

    OUT.mkdir(parents=True, exist_ok=True)
    current_path = OUT / "admin_dongs_incheon_20260701.csv"
    with current_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(current[0]))
        writer.writeheader()
        writer.writerows(sorted(current, key=lambda r: r["admin_dong_code_20260701"]))

    bridge_path = OUT / "admin_dong_crosswalk_20260630_to_20260701.csv"
    with bridge_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(bridge[0]))
        writer.writeheader()
        writer.writerows(sorted(bridge, key=lambda r: r["admin_dong_code_20260701"]))

    summary = {
        "current_dong_count": len(current),
        "pre_reform_dong_count": len(previous),
        "bridge_row_count": len(bridge),
        "bridge_unique_current_code_count": len({r["admin_dong_code_20260701"] for r in bridge}),
        "district_counts_20260701": dict(sorted(Counter(r["district_name_20260701"] for r in current).items())),
        "newly_created_current_code_count": sum(r["created_date"] == "20260701" for r in current),
        "important_limit": "This table maps attributes/codes only. It is not administrative geometry.",
    }
    (OUT / "admin_crosswalk_validation.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


if __name__ == "__main__":
    main()
