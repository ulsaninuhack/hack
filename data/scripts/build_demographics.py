#!/usr/bin/env python3
"""Build the 2026-07 Incheon administrative-dong demographic table.

The population source is dated 2026-06-30, immediately before Incheon's
2026-07-01 district reorganisation.  Its 162 rows are re-keyed to the active
2026-07-01 Ministry of the Interior and Safety (MOIS) administrative-dong
codes by exact, unique dong name.  The two household sources are already dated
2026-07-31 and must match the current MOIS keys directly.

Only aggregate resident-registration data are emitted.  No person-level data
are read or produced.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Iterable, Mapping, Sequence
from zipfile import ZipFile

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - environment-dependent guard
    raise SystemExit(
        "openpyxl is required to read the official MOIS code workbook. "
        "Install it with: python3 -m pip install openpyxl"
    ) from exc


INCHON = "인천광역시"
EXPECTED_DONGS = 162
CURRENT_CODE_MEMBER = "jscode20260701/KIKcd_H.20260701.xlsx"

POPULATION_FILENAME = "korea_population_by_sex_age_admin_dong_20260630.csv"
HOUSEHOLD_SIZE_FILENAME = "korea_households_by_size_admin_dong_20260731.csv"
ONE_PERSON_AGE_FILENAME = (
    "korea_one_person_households_by_sex_age_admin_dong_20260731.csv"
)
MOIS_CODE_FILENAME = "mois_admin_codes_20260701.zip"

DETAIL_OUTPUT = "demographics_admin_dong_202607.csv"
SUMMARY_OUTPUT = "demographics_district_summary_202607.csv"
VALIDATION_OUTPUT = "demographics_validation_202607.json"

AGE_COLUMN_PATTERN = re.compile(r"^(?P<age>\d+)세(?P<sex>남자|여자)$")
OPEN_AGE_COLUMN_PATTERN = re.compile(r"^(?P<age>\d+)세이상\s+(?P<sex>남자|여자)$")


class ValidationError(RuntimeError):
    """Raised when an input violates a required data contract."""


@dataclass(frozen=True)
class CurrentDong:
    admin_code: str
    district: str
    dong: str


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    script_path = Path(__file__).resolve()
    default_data_root = script_path.parents[1]

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--data-root",
        type=Path,
        default=default_data_root,
        help="i5-data directory containing raw/ and processed/",
    )
    return parser.parse_args(argv)


def normalize_code(value: object) -> str:
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    if not re.fullmatch(r"\d{10}", text):
        raise ValidationError(f"invalid 10-digit administrative code: {value!r}")
    return text


def normalize_dong_name_for_join(value: object) -> str:
    """Normalize known display-only punctuation without weakening identity.

    MOIS uses a middle dot in ``송림3·5동`` while resident-registration CSVs
    render the same official dong as ``송림3.5동``.  Normalizing that separator
    plus Unicode width and whitespace makes the cross-source join explicit and
    deterministic; district/name/code uniqueness checks still guard ambiguity.
    """

    text = unicodedata.normalize("NFKC", str(value)).strip().replace("·", ".")
    return re.sub(r"\s+", "", text)


def parse_int(value: object, *, field: str, allow_empty: bool = False) -> int:
    text = "" if value is None else str(value).strip().replace(",", "")
    if text == "" and allow_empty:
        return 0
    if not re.fullmatch(r"-?\d+", text):
        raise ValidationError(f"non-integer value in {field}: {value!r}")
    number = int(text)
    if number < 0:
        raise ValidationError(f"negative value in {field}: {number}")
    return number


def read_csv_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if not path.is_file():
        raise ValidationError(f"missing input file: {path}")
    with path.open("r", encoding="cp949", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ValidationError(f"CSV has no header: {path}")
        fieldnames = [name.strip() for name in reader.fieldnames]
        if fieldnames != reader.fieldnames:
            raise ValidationError(f"unexpected whitespace in CSV header: {path}")
        rows = [dict(row) for row in reader]
    return fieldnames, rows


def require_columns(path: Path, headers: Iterable[str], required: Iterable[str]) -> None:
    missing = sorted(set(required) - set(headers))
    if missing:
        raise ValidationError(f"{path.name} is missing columns: {missing}")


def incheon_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    headers, rows = read_csv_rows(path)
    require_columns(
        path,
        headers,
        ("행정기관코드", "기준연월", "시도명", "시군구명", "읍면동명"),
    )
    filtered = [row for row in rows if row["시도명"].strip() == INCHON]
    if len(filtered) != EXPECTED_DONGS:
        raise ValidationError(
            f"{path.name}: expected {EXPECTED_DONGS} Incheon rows, got {len(filtered)}"
        )
    return headers, filtered


def read_current_mois_dongs(path: Path) -> list[CurrentDong]:
    if not path.is_file():
        raise ValidationError(f"missing MOIS code archive: {path}")

    with ZipFile(path) as archive:
        try:
            workbook_bytes = archive.read(CURRENT_CODE_MEMBER)
        except KeyError as exc:
            raise ValidationError(
                f"{path.name} does not contain {CURRENT_CODE_MEMBER}"
            ) from exc

    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    row_iter = worksheet.iter_rows(values_only=True)
    headers = next(row_iter, None)
    expected_headers = ("행정동코드", "시도명", "시군구명", "읍면동명", "생성일자", "말소일자")
    if headers != expected_headers:
        raise ValidationError(
            f"unexpected MOIS workbook columns: {headers!r}; expected {expected_headers!r}"
        )

    current: list[CurrentDong] = []
    for code, province, district, dong, _created_at, deleted_at in row_iter:
        if province != INCHON or dong is None or deleted_at not in (None, ""):
            continue
        current.append(
            CurrentDong(
                admin_code=normalize_code(code),
                district=str(district).strip(),
                dong=str(dong).strip(),
            )
        )

    if len(current) != EXPECTED_DONGS:
        raise ValidationError(
            f"MOIS active-code list: expected {EXPECTED_DONGS} Incheon dongs, "
            f"got {len(current)}"
        )
    assert_unique(current, key=lambda row: row.admin_code, label="MOIS admin code")
    assert_unique(
        current,
        key=lambda row: normalize_dong_name_for_join(row.dong),
        label="normalized MOIS Incheon dong name",
    )
    return sorted(current, key=lambda row: row.admin_code)


def assert_unique(items: Iterable[object], *, key, label: str) -> None:
    values = [key(item) for item in items]
    duplicates = sorted(value for value, count in Counter(values).items() if count > 1)
    if duplicates:
        raise ValidationError(f"duplicate {label}: {duplicates}")


def require_single_reference_date(
    path: Path, rows: Sequence[Mapping[str, str]], expected_date: str
) -> str:
    dates = {row["기준연월"].strip() for row in rows}
    if dates != {expected_date}:
        raise ValidationError(
            f"{path.name}: expected reference date {expected_date}, got {sorted(dates)}"
        )
    return expected_date


def age_columns(headers: Iterable[str], *, minimum_age: int) -> list[str]:
    selected: list[tuple[int, str, str]] = []
    for column in headers:
        match = AGE_COLUMN_PATTERN.fullmatch(column)
        if match:
            age = int(match.group("age"))
            if age >= minimum_age:
                selected.append((age, match.group("sex"), column))
            continue
        match = OPEN_AGE_COLUMN_PATTERN.fullmatch(column)
        if match:
            age = int(match.group("age"))
            if age >= minimum_age:
                selected.append((age, match.group("sex"), column))

    selected.sort(key=lambda item: (item[1], item[0]))
    columns = [column for _age, _sex, column in selected]
    expected_count = (110 - minimum_age + 1) * 2
    if len(columns) != expected_count:
        raise ValidationError(
            f"expected {expected_count} sex-by-age columns for age {minimum_age}+, "
            f"got {len(columns)}"
        )
    return columns


def index_rows(
    rows: Sequence[dict[str, str]], *, key_field: str, label: str
) -> dict[str, dict[str, str]]:
    keys = [row[key_field].strip() for row in rows]
    duplicates = sorted(value for value, count in Counter(keys).items() if count > 1)
    if duplicates:
        raise ValidationError(f"duplicate {label}: {duplicates}")
    return {row[key_field].strip(): row for row in rows}


def validate_age_row(
    row: Mapping[str, str], headers: Iterable[str], *, source_name: str
) -> None:
    sex_total = parse_int(row["남자"], field=f"{source_name}.남자") + parse_int(
        row["여자"], field=f"{source_name}.여자"
    )
    total = parse_int(row["계"], field=f"{source_name}.계")
    if sex_total != total:
        raise ValidationError(
            f"{source_name} {row['읍면동명']}: 남자+여자={sex_total}, 계={total}"
        )

    all_age_columns = age_columns(headers, minimum_age=0)
    age_total = sum(
        parse_int(row[column], field=f"{source_name}.{column}", allow_empty=True)
        for column in all_age_columns
    )
    if age_total != total:
        raise ValidationError(
            f"{source_name} {row['읍면동명']}: age sum={age_total}, 계={total}"
        )


def sum_columns(row: Mapping[str, str], columns: Iterable[str], *, label: str) -> int:
    return sum(
        parse_int(row[column], field=f"{label}.{column}", allow_empty=True)
        for column in columns
    )


def safe_rate(numerator: int, denominator: int) -> str:
    if denominator == 0:
        return ""
    return f"{numerator / denominator:.6f}"


def build_detail_rows(
    current_dongs: Sequence[CurrentDong],
    population_headers: Sequence[str],
    population_rows: Sequence[dict[str, str]],
    household_headers: Sequence[str],
    household_rows: Sequence[dict[str, str]],
    one_person_headers: Sequence[str],
    one_person_rows: Sequence[dict[str, str]],
) -> tuple[list[dict[str, object]], dict[str, object]]:
    current_by_name = {
        normalize_dong_name_for_join(row.dong): row for row in current_dongs
    }
    current_by_code = {row.admin_code: row for row in current_dongs}

    assert_unique(
        population_rows,
        key=lambda row: normalize_dong_name_for_join(row["읍면동명"]),
        label="normalized population dong",
    )
    population_names = {
        normalize_dong_name_for_join(row["읍면동명"]) for row in population_rows
    }
    current_names = set(current_by_name)
    if population_names != current_names:
        raise ValidationError(
            "population/current MOIS dong-name sets differ: "
            f"population_only={sorted(population_names - current_names)}, "
            f"mois_only={sorted(current_names - population_names)}"
        )
    population_by_current_code: dict[str, dict[str, str]] = {}
    for row in population_rows:
        current_code = current_by_name[
            normalize_dong_name_for_join(row["읍면동명"])
        ].admin_code
        if current_code in population_by_current_code:
            raise ValidationError(f"multiple population rows map to {current_code}")
        population_by_current_code[current_code] = row

    household_by_code = index_rows(
        household_rows, key_field="행정기관코드", label="household admin code"
    )
    one_person_by_code = index_rows(
        one_person_rows, key_field="행정기관코드", label="one-person admin code"
    )
    if set(household_by_code) != set(current_by_code):
        raise ValidationError("household-size codes do not exactly match current MOIS codes")
    if set(one_person_by_code) != set(current_by_code):
        raise ValidationError("one-person-by-age codes do not exactly match current MOIS codes")

    population_65_columns = age_columns(population_headers, minimum_age=65)
    one_person_65_columns = age_columns(one_person_headers, minimum_age=65)
    household_size_columns = [
        "1인세대",
        "2인세대",
        "3인세대",
        "4인세대",
        "5인세대",
        "6인세대",
        "7인세대",
        "8인세대",
        "9인세대",
        "10인이상세대",
    ]
    require_columns(Path(HOUSEHOLD_SIZE_FILENAME), household_headers, household_size_columns)

    detail_rows: list[dict[str, object]] = []
    changed_codes = 0
    changed_districts = 0
    for current in current_dongs:
        population = population_by_current_code[current.admin_code]
        household = household_by_code[current.admin_code]
        one_person = one_person_by_code[current.admin_code]

        if (
            household["시군구명"].strip() != current.district
            or normalize_dong_name_for_join(household["읍면동명"])
            != normalize_dong_name_for_join(current.dong)
        ):
            raise ValidationError(f"household key/name mismatch for {current.admin_code}")
        if (
            one_person["시군구명"].strip() != current.district
            or normalize_dong_name_for_join(one_person["읍면동명"])
            != normalize_dong_name_for_join(current.dong)
        ):
            raise ValidationError(f"one-person key/name mismatch for {current.admin_code}")

        validate_age_row(population, population_headers, source_name="population")
        validate_age_row(one_person, one_person_headers, source_name="one-person")

        total_population = parse_int(population["계"], field="population.계")
        population_65_plus = sum_columns(
            population, population_65_columns, label="population"
        )
        total_households = parse_int(
            household["전체세대수"], field="household.전체세대수"
        )
        household_sum = sum_columns(
            household, household_size_columns, label="household"
        )
        if household_sum != total_households:
            raise ValidationError(
                f"{current.dong}: household-size sum={household_sum}, "
                f"전체세대수={total_households}"
            )
        one_person_households = parse_int(
            household["1인세대"], field="household.1인세대"
        )
        one_person_total_from_age = parse_int(
            one_person["계"], field="one-person.계"
        )
        if one_person_households != one_person_total_from_age:
            raise ValidationError(
                f"{current.dong}: household 1인세대={one_person_households}, "
                f"age-table 계={one_person_total_from_age}"
            )
        one_person_65_plus = sum_columns(
            one_person, one_person_65_columns, label="one-person"
        )

        source_code = normalize_code(population["행정기관코드"])
        source_district = population["시군구명"].strip()
        code_changed = source_code != current.admin_code
        district_changed = source_district != current.district
        changed_codes += int(code_changed)
        changed_districts += int(district_changed)

        detail_rows.append(
            {
                "admin_dong_code_20260701": current.admin_code,
                "province_name": INCHON,
                "district_name_20260701": current.district,
                "admin_dong_name": current.dong,
                "population_reference_date": population["기준연월"].strip(),
                "household_reference_date": household["기준연월"].strip(),
                "population_source_admin_code_20260630": source_code,
                "population_source_district_name_20260630": source_district,
                "population_code_changed_at_reform": str(code_changed).lower(),
                "population_district_changed_at_reform": str(district_changed).lower(),
                "total_population": total_population,
                "population_age_65_plus": population_65_plus,
                "total_households": total_households,
                "one_person_households": one_person_households,
                "one_person_households_age_65_plus": one_person_65_plus,
                "population_age_65_plus_share": safe_rate(
                    population_65_plus, total_population
                ),
                "one_person_household_share": safe_rate(
                    one_person_households, total_households
                ),
                "age_65_plus_one_person_share_of_age_65_plus_population": safe_rate(
                    one_person_65_plus, population_65_plus
                ),
            }
        )

    if len(detail_rows) != EXPECTED_DONGS:
        raise ValidationError(f"expected {EXPECTED_DONGS} detail rows, got {len(detail_rows)}")

    validation = {
        "row_count": len(detail_rows),
        "unique_current_admin_codes": len(
            {row["admin_dong_code_20260701"] for row in detail_rows}
        ),
        "unique_current_dong_names": len({row["admin_dong_name"] for row in detail_rows}),
        "population_rows_rekeyed_to_new_admin_code": changed_codes,
        "population_rows_reassigned_to_new_district_name": changed_districts,
        "district_count": len({row["district_name_20260701"] for row in detail_rows}),
        "age_65_plus_column_count_per_age_table": len(population_65_columns),
        "checks": {
            "mois_current_keys_are_unique": True,
            "population_dong_names_match_current_mois_names": True,
            "july_household_keys_match_current_mois_keys": True,
            "july_one_person_keys_match_current_mois_keys": True,
            "population_sex_and_age_totals_reconcile": True,
            "one_person_sex_and_age_totals_reconcile": True,
            "household_size_totals_reconcile": True,
            "one_person_totals_match_across_household_sources": True,
        },
    }
    return detail_rows, validation


def build_district_summary(detail_rows: Sequence[Mapping[str, object]]) -> list[dict[str, object]]:
    grouped: dict[str, list[Mapping[str, object]]] = defaultdict(list)
    for row in detail_rows:
        grouped[str(row["district_name_20260701"])].append(row)

    result: list[dict[str, object]] = []
    for district in sorted(grouped):
        rows = grouped[district]
        total_population = sum(int(row["total_population"]) for row in rows)
        population_65_plus = sum(int(row["population_age_65_plus"]) for row in rows)
        total_households = sum(int(row["total_households"]) for row in rows)
        one_person_households = sum(int(row["one_person_households"]) for row in rows)
        one_person_65_plus = sum(
            int(row["one_person_households_age_65_plus"]) for row in rows
        )
        result.append(
            {
                "province_name": INCHON,
                "district_name_20260701": district,
                "admin_dong_count": len(rows),
                "total_population": total_population,
                "population_age_65_plus": population_65_plus,
                "total_households": total_households,
                "one_person_households": one_person_households,
                "one_person_households_age_65_plus": one_person_65_plus,
                "population_age_65_plus_share": safe_rate(
                    population_65_plus, total_population
                ),
                "one_person_household_share": safe_rate(
                    one_person_households, total_households
                ),
                "age_65_plus_one_person_share_of_age_65_plus_population": safe_rate(
                    one_person_65_plus, population_65_plus
                ),
            }
        )
    if len(result) != 11:
        raise ValidationError(f"expected 11 current Incheon districts, got {len(result)}")
    return result


def write_csv(path: Path, rows: Sequence[Mapping[str, object]]) -> None:
    if not rows:
        raise ValidationError(f"refusing to write empty output: {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    data_root = args.data_root.resolve()
    demographics_dir = data_root / "raw" / "demographics"
    boundaries_dir = data_root / "raw" / "boundaries"
    output_dir = data_root / "processed"

    population_path = demographics_dir / POPULATION_FILENAME
    household_path = demographics_dir / HOUSEHOLD_SIZE_FILENAME
    one_person_path = demographics_dir / ONE_PERSON_AGE_FILENAME
    mois_path = boundaries_dir / MOIS_CODE_FILENAME

    population_headers, population_rows = incheon_rows(population_path)
    household_headers, household_rows = incheon_rows(household_path)
    one_person_headers, one_person_rows = incheon_rows(one_person_path)
    require_columns(population_path, population_headers, ("계", "남자", "여자"))
    require_columns(household_path, household_headers, ("전체세대수", "1인세대"))
    require_columns(one_person_path, one_person_headers, ("계", "남자", "여자"))

    population_date = require_single_reference_date(
        population_path, population_rows, "2026-06-30"
    )
    household_date = require_single_reference_date(
        household_path, household_rows, "2026-07-31"
    )
    one_person_date = require_single_reference_date(
        one_person_path, one_person_rows, "2026-07-31"
    )
    current_dongs = read_current_mois_dongs(mois_path)

    detail_rows, validation = build_detail_rows(
        current_dongs,
        population_headers,
        population_rows,
        household_headers,
        household_rows,
        one_person_headers,
        one_person_rows,
    )
    summary_rows = build_district_summary(detail_rows)

    detail_path = output_dir / DETAIL_OUTPUT
    summary_path = output_dir / SUMMARY_OUTPUT
    validation_path = output_dir / VALIDATION_OUTPUT
    write_csv(detail_path, detail_rows)
    write_csv(summary_path, summary_rows)

    validation.update(
        {
            "population_reference_date": population_date,
            "household_reference_date": household_date,
            "one_person_age_reference_date": one_person_date,
            "current_key_reference_date": "2026-07-01",
            "inputs": {
                "population": str(population_path.relative_to(data_root)),
                "household_size": str(household_path.relative_to(data_root)),
                "one_person_by_sex_age": str(one_person_path.relative_to(data_root)),
                "mois_admin_codes": str(mois_path.relative_to(data_root)),
            },
            "outputs": {
                "admin_dong": str(detail_path.relative_to(data_root)),
                "district_summary": str(summary_path.relative_to(data_root)),
            },
        }
    )
    validation_path.write_text(
        json.dumps(validation, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(
        f"Built {len(detail_rows)} dongs across {len(summary_rows)} districts: "
        f"{detail_path}"
    )
    print(
        "Population rows re-keyed at the 2026-07-01 reform: "
        f"{validation['population_rows_rekeyed_to_new_admin_code']}"
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ValidationError as exc:
        print(f"validation error: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
